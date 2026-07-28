"""Le os relatorios .xlsx exportados do painel de usinas (todas as abas)."""
import re
from collections import Counter

import openpyxl

_TRANS = str.maketrans("áàâãäéèêëíìîïóòôõöúùûüçñÁÀÂÃÄÉÈÊËÍÌÎÏÓÒÔÕÖÚÙÛÜÇÑ",
                        "aaaaaeeeeiiiiooooouuuucnAAAAAEEEEIIIIOOOOOUUUUCN")

_MES_ABREV = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}


def _norm(v):
    s = str(v if v is not None else "").strip().lower()
    return s.translate(_TRANS)


def _texto(v):
    if v is None:
        return ""
    return str(v).strip()


def _num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("\xa0", "")
    s = re.sub(r"[^\d,.\-]", "", s)
    if not s or s == "-":
        return 0.0
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _pct(v):
    return _num(v) / 100.0


def _mes_ref_iso(valor):
    """'MM/YYYY' ou 'YYYY-MM' -> 'YYYY-MM'."""
    s = _texto(valor)
    m = re.match(r"^(\d{4})-(\d{1,2})$", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    m = re.match(r"^(\d{1,2})/(\d{4})$", s)
    if m:
        mes, ano = int(m.group(1)), int(m.group(2))
        if 1 <= mes <= 12:
            return f"{ano:04d}-{mes:02d}"
    return None


def parse_periodo_livre(texto):
    """Extrai 'YYYY-MM' de textos como 'Jun/2026' ou 'Junho / 2026'."""
    s = _norm(texto)
    m = re.search(r"([a-z]{3,9})\s*/?\s*(\d{4})", s)
    if not m:
        return None
    mes = _MES_ABREV.get(m.group(1)[:3])
    if not mes:
        return None
    return f"{int(m.group(2)):04d}-{mes:02d}"


def _usina_selecionada(texto):
    m = re.search(r"usina:\s*(.+?)\s*\|", texto, re.IGNORECASE)
    return m.group(1).strip() if m else None


def _find_header(ws, obrigatorios, max_scan=25):
    """Acha a linha de cabecalho: primeira linha cujas celulas (normalizadas)
    contem todas as substrings obrigatorias. Retorna (numero_linha, {header_norm: idx})."""
    for row in ws.iter_rows(min_row=1, max_row=min(max_scan, ws.max_row)):
        cells = [_norm(c.value) for c in row]
        if all(any(req in c for c in cells if c) for req in obrigatorios):
            col_map = {}
            for idx, c in enumerate(cells):
                if c:
                    col_map[c] = idx
            return row[0].row, col_map
    return None, None


def _col(col_map, *substrings):
    for key, idx in col_map.items():
        if all(s in key for s in substrings):
            return idx
    return None


def _cell(row, idx):
    return row[idx] if idx is not None and idx < len(row) else None


_UC_LABEL_RE = re.compile(r"^uc\b", re.IGNORECASE)


def _concessionaria_valida(v):
    """Alguns exports trazem o rotulo da UC ('UC 123-4') na coluna
    'Concessionaria' em vez do nome da distribuidora. Descarta esse lixo."""
    s = _texto(v)
    if _UC_LABEL_RE.match(s):
        return ""
    return s


# ── Aba "Unidades Consumidoras" ──────────────────────────────────────────────

def parse_unidades_consumidoras(ws):
    linha, col_map = _find_header(ws, ["uc", "status", "usinas geradoras"])
    if linha is None:
        raise ValueError("Aba 'Unidades Consumidoras': cabecalho nao encontrado.")

    c_id       = _col(col_map, "id fatura")
    c_conc     = _col(col_map, "concession")
    c_nome     = _col(col_map, "nome")
    c_cod      = _col(col_map, "codigo da venda")
    c_uc       = _col(col_map, "uc")
    c_mes      = _col(col_map, "mes referencia")
    c_consumo  = _col(col_map, "consumo real")
    c_comp     = _col(col_map, "energia compensada")
    c_usina    = _col(col_map, "usinas geradoras")
    c_emissao  = _col(col_map, "emissao")
    c_venc     = _col(col_map, "vencimento")
    c_vconc    = _col(col_map, "valor concessionaria")
    c_vreceber = _col(col_map, "valor a receber")
    c_vfat     = _col(col_map, "valor faturado")
    c_datapag  = _col(col_map, "data do pagamento")
    c_status   = _col(col_map, "status")

    registros = []
    periodos = Counter()
    for row in ws.iter_rows(min_row=linha + 1, values_only=True):
        uc = _texto(_cell(row, c_uc))
        if not uc or _norm(uc) == "uc":
            continue
        mes_ref = _mes_ref_iso(_cell(row, c_mes))
        if not mes_ref:
            continue
        periodos[mes_ref] += 1

        id_fatura = _texto(_cell(row, c_id))
        registros.append({
            "id_fatura":              id_fatura if id_fatura not in ("", "-") else None,
            "uc":                     uc,
            "nome_cliente":           _texto(_cell(row, c_nome)),
            "codigo_venda":           _texto(_cell(row, c_cod)),
            "concessionaria":         _concessionaria_valida(_cell(row, c_conc)),
            "mes_referencia":         mes_ref,
            "consumo_real_kwh":       _num(_cell(row, c_consumo)),
            "energia_compensada_kwh": _num(_cell(row, c_comp)),
            "usina":                  _texto(_cell(row, c_usina)),
            "emissao":                _texto(_cell(row, c_emissao)),
            "vencimento":             _texto(_cell(row, c_venc)),
            "valor_concessionaria":   _num(_cell(row, c_vconc)),
            "valor_a_receber":        _num(_cell(row, c_vreceber)),
            "valor_faturado":         _num(_cell(row, c_vfat)) if c_vfat is not None else None,
            "data_pagamento":         _texto(_cell(row, c_datapag)),
            "status":                 _texto(_cell(row, c_status)),
        })

    if not registros:
        raise ValueError("Aba 'Unidades Consumidoras': nenhum registro valido encontrado.")
    periodo_dominante = periodos.most_common(1)[0][0]
    return registros, periodo_dominante


# ── Aba "Rateios Cadastrados" ────────────────────────────────────────────────

def parse_rateios(ws):
    linha, col_map = _find_header(ws, ["unidade consumidora", "usina", "porcentagem"])
    if linha is None:
        return []

    c_uc     = _col(col_map, "unidade consumidora")
    c_usina  = _col(col_map, "usina")
    c_cod    = _col(col_map, "codigo venda")
    c_mes    = _col(col_map, "mes de referencia")
    c_inj    = _col(col_map, "injecao prevista")
    c_pct    = _col(col_map, "porcentagem")
    c_tipo   = _col(col_map, "tipo gd")
    c_stpag  = _col(col_map, "status pagamento")
    c_idst   = _col(col_map, "id status")
    c_ultmes = _col(col_map, "ultimo mes")
    c_ultval = _col(col_map, "ultimo valor pago")

    out = []
    for row in ws.iter_rows(min_row=linha + 1, values_only=True):
        uc = _texto(_cell(row, c_uc))
        if not uc:
            continue
        mes_ref = _mes_ref_iso(_cell(row, c_mes))
        if not mes_ref:
            continue
        out.append({
            "uc":                    uc,
            "usina":                 _texto(_cell(row, c_usina)),
            "codigo_venda":          _texto(_cell(row, c_cod)),
            "mes_referencia":        mes_ref,
            "injecao_prevista_kwh":  _num(_cell(row, c_inj)),
            "porcentagem":           _num(_cell(row, c_pct)),
            "tipo_gd":               _texto(_cell(row, c_tipo)),
            "status_pagamento":      _texto(_cell(row, c_stpag)),
            "id_status":             _texto(_cell(row, c_idst)),
            "ultimo_mes_referencia": _mes_ref_iso(_cell(row, c_ultmes)) or "",
            "ultimo_valor_pago":     _num(_cell(row, c_ultval)),
        })
    return out


# ── Aba "Analise por Beneficiaria" ──────────────────────────────────────────

def parse_analise_beneficiaria(ws, periodo_snapshot):
    linha, col_map = _find_header(ws, ["cod", "venda", "inadimpl"])
    if linha is None:
        return []

    c_cod    = _col(col_map, "venda")
    c_uc     = _col(col_map, "uc")
    c_prev   = _col(col_map, "prevista consumo")
    c_real   = _col(col_map, "real consumo")
    c_pctrp  = _col(col_map, "consumo real")
    c_comp   = _col(col_map, "energia compensada")
    c_pctcc  = _col(col_map, "comp", "consumo")
    c_qtd    = _col(col_map, "qtd faturas")
    c_pagas  = _col(col_map, "pagas")
    c_atras  = _col(col_map, "atrasadas")
    c_avenc  = _col(col_map, "a vencer")
    c_volpag = _col(col_map, "energia paga")
    c_volina = _col(col_map, "energia inadimpl")
    c_volave = _col(col_map, "energia a vencer")
    c_inad   = _col(col_map, "inadimpl", "%")

    out = []
    for row in ws.iter_rows(min_row=linha + 1, values_only=True):
        uc = _texto(_cell(row, c_uc))
        if not uc:
            continue
        out.append({
            "uc":                       uc,
            "codigo_venda":             _texto(_cell(row, c_cod)),
            "periodo_snapshot":         periodo_snapshot,
            "media_prevista_consumo":   _num(_cell(row, c_prev)),
            "media_real_consumo":       _num(_cell(row, c_real)),
            "pct_consumo_real_x_prev":  _pct(_cell(row, c_pctrp)),
            "media_energia_compensada": _num(_cell(row, c_comp)),
            "pct_comp_x_consumo":       _pct(_cell(row, c_pctcc)),
            "qtd_faturas":              int(_num(_cell(row, c_qtd))),
            "pagas":                    int(_num(_cell(row, c_pagas))),
            "atrasadas":                int(_num(_cell(row, c_atras))),
            "a_vencer":                 int(_num(_cell(row, c_avenc))),
            "vol_energia_paga":         _num(_cell(row, c_volpag)),
            "vol_energia_inadimplente": _num(_cell(row, c_volina)),
            "vol_energia_a_vencer":     _num(_cell(row, c_volave)),
            "inadimplencia_pct":        _pct(_cell(row, c_inad)),
        })
    return out


# ── Abas "Dashboard Painel do Gerador" / "Relatorio de Performance" ─────────

def parse_grade_mensal(ws, periodo_export, fonte):
    """Extrai a grade 'Indicador | Jan..Dez | Total' e devolve valores por mes."""
    linha, col_map = _find_header(ws, ["indicador", "jan", "dez"])
    if linha is None:
        return [], None

    usina_sel = None
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        for v in row:
            if v and "usina:" in str(v).lower():
                usina_sel = _usina_selecionada(str(v))
                break
        if usina_sel:
            break
    usina_sel = usina_sel or "Todas as usinas"

    ano = int(periodo_export[:4]) if periodo_export else None
    meses_idx = {}
    for key, idx in col_map.items():
        if key in _MES_ABREV:
            meses_idx[_MES_ABREV[key]] = idx

    saida = []
    for row in ws.iter_rows(min_row=linha + 1, values_only=True):
        indicador = _texto(row[0]) if row else ""
        if not indicador:
            continue
        if _norm(indicador) == "indicador":
            continue
        for mes_num, idx in meses_idx.items():
            if ano is None:
                continue
            valor = _cell(row, idx)
            if valor is None or valor == "":
                continue
            saida.append({
                "usina_selecionada": usina_sel,
                "periodo_export":    periodo_export,
                "fonte":             fonte,
                "indicador":         indicador,
                "mes_referencia":    f"{ano:04d}-{mes_num:02d}",
                "valor":             _num(valor),
            })
    return saida, usina_sel


def parse_performance_kpis(ws, periodo_export):
    """Extrai os KPIs de cabecalho da aba 'Relatorio de Performance'."""
    usina_sel = None
    kpis = {}
    labels_desejados = {
        "id usina": "id_usina",
        "instalacao": "instalacao",
        "concession": "concessionaria",
        "potencia conexao": "potencia_conexao",
        "potencia instalada": "potencia_instalada",
        "fonte": "fonte",
        "geracao prevista acum": "geracao_prevista_acum_mwh",
        "geracao real acum": "geracao_real_acum_mwh",
        "valor faturado": "valor_faturado",
        "valor pago": "valor_pago",
        "valor a vencer": "valor_a_vencer",
        "valor atrasado": "valor_atrasado",
        "inadimpl": "inadimplencia_pct",
        "receita liquida prevista": "receita_liquida_prevista",
        "receita liquida realizada": "receita_liquida_realizada",
    }

    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        for v in row:
            if v and "usina:" in str(v).lower():
                usina_sel = _usina_selecionada(str(v))

        for i, v in enumerate(row):
            norm = _norm(v)
            if not norm:
                continue
            for chave, campo in labels_desejados.items():
                if chave in norm:
                    valor = row[i + 1] if i + 1 < len(row) else None
                    if valor in (None, ""):
                        continue
                    if campo == "inadimplencia_pct":
                        kpis[campo] = _pct(valor)
                    elif campo in ("id_usina", "instalacao", "concessionaria",
                                   "potencia_conexao", "potencia_instalada", "fonte"):
                        kpis[campo] = _texto(valor)
                    else:
                        kpis[campo] = _num(valor)
                    break

    return usina_sel or "Todas as usinas", kpis


# ── Orquestrador ─────────────────────────────────────────────────────────────

def parse_relatorio_completo(caminho_ou_stream):
    wb = openpyxl.load_workbook(caminho_ou_stream, data_only=True)
    resultado = {
        "faturas": [], "periodo_dominante": None,
        "rateios": [], "analise_beneficiaria": [],
        "portfolio_mensal": [], "portfolio_kpis": None,
    }

    if "Unidades Consumidoras" in wb.sheetnames:
        registros, periodo = parse_unidades_consumidoras(wb["Unidades Consumidoras"])
        resultado["faturas"] = registros
        resultado["periodo_dominante"] = periodo

    periodo_export = resultado["periodo_dominante"]

    if "Rateios Cadastrados" in wb.sheetnames:
        resultado["rateios"] = parse_rateios(wb["Rateios Cadastrados"])

    if "Análise por Beneficiária" in wb.sheetnames or "Analise por Beneficiaria" in wb.sheetnames:
        aba = wb["Análise por Beneficiária"] if "Análise por Beneficiária" in wb.sheetnames \
            else wb["Analise por Beneficiaria"]
        resultado["analise_beneficiaria"] = parse_analise_beneficiaria(aba, periodo_export)

    grade_total = []
    if "Dashboard Painel do Gerador" in wb.sheetnames:
        grade, _ = parse_grade_mensal(wb["Dashboard Painel do Gerador"], periodo_export, "dashboard_gerador")
        grade_total += grade
    nome_perf = "Relatório de Performance" if "Relatório de Performance" in wb.sheetnames else "Relatorio de Performance"
    if nome_perf in wb.sheetnames:
        grade, usina_sel = parse_grade_mensal(wb[nome_perf], periodo_export, "performance")
        grade_total += grade
        usina_sel_kpi, kpis = parse_performance_kpis(wb[nome_perf], periodo_export)
        if kpis:
            resultado["portfolio_kpis"] = {
                "usina_selecionada": usina_sel_kpi,
                "periodo_export": periodo_export,
                **kpis,
            }
    resultado["portfolio_mensal"] = grade_total

    return resultado
