"""
Importação única do histórico da planilha FUP Operações.xlsx pro banco do
painel. Depois de rodado, a planilha não é mais usada -- toda a operação
passa a ser feita pelo sistema web.

Uso: python importar_historico.py
"""
import re
import shutil
import tempfile
import unicodedata
from datetime import date, timedelta

import openpyxl

import db

CAMINHO_PLANILHA = (
    r"D:\Alexandria\OneDrive - Alexandria Industria de Geradores SA"
    r"\Diretoria de Operações - Documents\6. Ativos Próprios"
    r"\Gestão de Parceiros\FUP Operações.xlsx"
)

PALAVRAS_ENCERRAMENTO = ["RESCIS", "CANCELAD", "ENCERR"]


def sem_acento(texto):
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def normalizar_secao(nome):
    n = sem_acento(str(nome)).upper()
    n = re.sub(r"\(.*?\)", "", n)  # remove "(3 e 4)", "(AUGUSTO)" etc.
    n = re.sub(r"[^A-Z ]", " ", n)
    n = re.sub(r"\s+", " ", n).strip()
    return n


MAPA_ETAPAS = {
    "SEPARACAO E RATEIO": "Separando Clientes",
    "CONFIRMAR APROVACAO REPROVACAO DE RATEIO": "Aguardando Aprovação Rateio",
    "SOLICITAR TT USINA": "TT Usina",
    "FUP TT USINAS": "TT Usina",
    "CONFIRMACAO TT": "TT Usina",
    "CADASTRO RATEIO": "Aguardando Rateio",
    "CORRECAO RATEIO": "Refazer Rateio",
    "CONFIRMACAO APROVACAO RATEIO": "Aguardando Aprovação Rateio",
    "CADASTRAR TT": "TT Usina",
    "CONFIRMAR TT": "TT Usina",
    "CADASTRAR RATEIO": "Aguardando Rateio",
    "CONFIRMAR RATEIO": "Aguardando Aprovação Rateio",
    "SOLICITAR TT": "TT Usina",
    "CONFIRMAR TT USINA": "TT Usina",
    "CONFIRMAR TT CLIENTE": "TT Cliente",
    "SEPARAR CLIENTE": "Separando Clientes",
    "PROTOCOLAR RATEIO": "Aguardando Rateio",
    "ACOMPANHAR TT USINA": "TT Usina",
    "SEPARACAO PENDENTE": "Separando Clientes",
    "TT CLIENTE PENDENTE": "TT Cliente",
    "AGUARDANDO SEPARACAO DE CLIENTE": "Separando Clientes",
    "CLIENTES SEPARADOS": "TT Cliente",
    "TT CLIENTE EM ANDAMENTO": "TT Cliente",
    "FAZER RATEIO": "Aguardando Rateio",
    "TT CLIENTE": "TT Cliente",
    "APROVAR RATEIO": "Aguardando Aprovação Rateio",
    "REFAZER RATEIO": "Refazer Rateio",
    "REDISTRIBUICAO DE CREDITOS": "Aguardando Rateio",
    "TT USINA EM ANDAMENTO": "TT Usina",
    "SEM DISPONIBILIDADE DE CLIENTES": "Sem Clientes",
    "AGUARDANDO SEPARACAO DE CLIENTES": "Separando Clientes",
    "AGUARDANDO TT CLIENTE": "TT Cliente",
    "AGUARDANDO RATEIO": "Aguardando Rateio",
    "AGUARDANDO APROV RATEIO": "Aguardando Aprovação Rateio",
    "TT USINA": "TT Usina",
    "SEM DISP CLIENTE": "Sem Clientes",
    "AGUARDANDO SEPARACAO CLIENTE": "Separando Clientes",
    "AGUARD TT CLIENTE": "TT Cliente",
    "AGUARDANDO APROVACAO RATEIO": "Aguardando Aprovação Rateio",
    "ASSINADO COM PENDENCIA": "Assinado c/ Pendência",
    "ASSINADOS COM PENDENCIA": "Assinado c/ Pendência",
    "SEM CLIENTE": "Sem Clientes",
    "SEM CLIENTES": "Sem Clientes",
    "SEPARANDO CLIENTES": "Separando Clientes",
    "CONSUMO DE SALDO ACUMULADO": "Consumo de Saldo Acumulado",
    "TT CLIENTE DE ANDAMENTO": "TT Cliente",
}


def mapear_etapa(nome_secao, nao_mapeadas):
    chave = normalizar_secao(nome_secao)
    if chave in MAPA_ETAPAS:
        return MAPA_ETAPAS[chave]
    nao_mapeadas.add(nome_secao)
    return "TT Usina"  # fallback conservador


def abrir_planilha():
    try:
        return openpyxl.load_workbook(CAMINHO_PLANILHA, data_only=True)
    except PermissionError:
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        shutil.copy2(CAMINHO_PLANILHA, tmp.name)
        return openpyxl.load_workbook(tmp.name, data_only=True)


def data_da_aba(nome_aba):
    nome = nome_aba.strip()
    m = re.match(r"^(\d{2})\.(\d{2})$", nome)
    if not m:
        return None
    dia, mes = int(m.group(1)), int(m.group(2))
    return date(2026, mes, dia)


HEADER_ALIASES = {
    "UG": ["UG"],
    "Nome UFV": ["Nome UFV"],
    "Concessionária": ["Concessionária", "Concessionaria"],
    "Dono Carteira": ["Dono Carteira", "Responsável", "Responsavel"],
    "Dt. Ass. Cont.": ["Dt. Ass. Cont."],
    "Observação": ["Observação", "Observacao"],
}


def construir_mapa_colunas(ws):
    for r in range(1, 4):
        valores = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(isinstance(v, str) and v.strip() == "UG" for v in valores):
            mapa = {}
            for idx, v in enumerate(valores, start=1):
                if isinstance(v, str) and v.strip():
                    mapa[v.strip()] = idx
            return r, mapa
    return None, None


def col(mapa, campo, max_col):
    for alias in HEADER_ALIASES[campo]:
        if alias in mapa:
            return mapa[alias]
    if campo == "Observação":
        return max_col
    return None


def ler_aba(ws, nao_mapeadas):
    header_row, mapa = construir_mapa_colunas(ws)
    if header_row is None:
        return {}
    ug_col = col(mapa, "UG", ws.max_column)
    nome_col = col(mapa, "Nome UFV", ws.max_column)
    conc_col = col(mapa, "Concessionária", ws.max_column)
    dono_col = col(mapa, "Dono Carteira", ws.max_column)
    dtass_col = col(mapa, "Dt. Ass. Cont.", ws.max_column)
    obs_col = col(mapa, "Observação", ws.max_column)

    registros = {}
    secao_atual = None
    for r in range(header_row, ws.max_row + 1):
        a_val = ws.cell(row=r, column=1).value
        ug_val = ws.cell(row=r, column=ug_col).value if ug_col else None

        eh_linha_de_cabecalho = isinstance(ug_val, str) and ug_val.strip().upper() == "UG"
        if eh_linha_de_cabecalho:
            if a_val:
                secao_atual = str(a_val).strip()
            continue

        if ug_val is None or str(ug_val).strip() == "":
            if a_val:
                secao_atual = str(a_val).strip()
            continue

        if a_val:
            secao_atual = str(a_val).strip()
        if secao_atual is None:
            continue

        ug_norm = db.normalizar_ug(ug_val)
        if not ug_norm:
            continue

        dt_ass = ws.cell(row=r, column=dtass_col).value if dtass_col else None
        registros[ug_norm] = {
            "ug_raw": str(ug_val).strip(),
            "nome_ufv": str(ws.cell(row=r, column=nome_col).value or "").strip() if nome_col else "",
            "concessionaria": str(ws.cell(row=r, column=conc_col).value or "").strip() if conc_col else "",
            "dono_carteira": str(ws.cell(row=r, column=dono_col).value or "").strip() if dono_col else "",
            "data_assinatura": dt_ass.date().isoformat() if hasattr(dt_ass, "date") else None,
            "observacao": str(ws.cell(row=r, column=obs_col).value or "").strip() if obs_col else "",
            "etapa": mapear_etapa(secao_atual, nao_mapeadas),
        }
    return registros


def contem_encerramento(observacoes):
    texto = sem_acento(" ".join(observacoes)).upper()
    return any(p in texto for p in PALAVRAS_ENCERRAMENTO)


def principal():
    wb = abrir_planilha()
    nao_mapeadas = set()

    abas_datadas = []
    for nome_aba in wb.sheetnames:
        d = data_da_aba(nome_aba)
        if d:
            abas_datadas.append((d, nome_aba))
    abas_datadas.sort(key=lambda x: x[0])

    if not abas_datadas:
        raise RuntimeError("Nenhuma aba com nome no formato dd.mm encontrada.")

    ultima_data = abas_datadas[-1][0]

    # timeline por usina: ug -> lista de (data, registro)
    timeline = {}
    for data_aba, nome_aba in abas_datadas:
        ws = wb[nome_aba]
        registros = ler_aba(ws, nao_mapeadas)
        for ug, reg in registros.items():
            timeline.setdefault(ug, []).append((data_aba, reg))

    datas_ordenadas = [d for d, _ in abas_datadas]

    db.iniciar_banco()
    with db.conectar() as conn:
        conn.execute("DELETE FROM historico_etapas")
        conn.execute("DELETE FROM usinas")

        n_ativas = n_operacao = n_rescindidas = 0

        for ug, aparicoes in timeline.items():
            aparicoes.sort(key=lambda x: x[0])
            attrs_final = aparicoes[-1][1]
            todas_obs = [a[1]["observacao"] for a in aparicoes if a[1]["observacao"]]

            data_assinatura = None
            for _, reg in reversed(aparicoes):
                if reg["data_assinatura"]:
                    data_assinatura = reg["data_assinatura"]
                    break

            segmentos = []
            etapa_seg = aparicoes[0][1]["etapa"]
            inicio_seg = aparicoes[0][0]
            for i in range(1, len(aparicoes)):
                data_i, reg_i = aparicoes[i]
                if reg_i["etapa"] != etapa_seg:
                    segmentos.append((etapa_seg, inicio_seg, data_i))
                    etapa_seg = reg_i["etapa"]
                    inicio_seg = data_i

            ultima_aparicao = aparicoes[-1][0]
            if ultima_aparicao == ultima_data:
                segmentos.append((etapa_seg, inicio_seg, None))
                status = "ativa"
                n_ativas += 1
            else:
                idx = datas_ordenadas.index(ultima_aparicao)
                data_saida = datas_ordenadas[idx + 1] if idx + 1 < len(datas_ordenadas) else ultima_aparicao + timedelta(days=7)
                segmentos.append((etapa_seg, inicio_seg, data_saida))
                if contem_encerramento(todas_obs):
                    status = "rescindida"
                    n_rescindidas += 1
                else:
                    status = "operacao"
                    n_operacao += 1

            etapa_atual_stage, etapa_atual_inicio, _ = segmentos[-1]

            cur = conn.execute(
                """INSERT INTO usinas
                   (ug, ug_raw, nome_ufv, concessionaria, dono_carteira,
                    data_assinatura_contrato, etapa_atual, data_entrada_etapa_atual,
                    status, observacao)
                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (
                    ug,
                    attrs_final["ug_raw"],
                    attrs_final["nome_ufv"],
                    attrs_final["concessionaria"],
                    attrs_final["dono_carteira"],
                    data_assinatura,
                    etapa_atual_stage,
                    etapa_atual_inicio.isoformat(),
                    status,
                    attrs_final["observacao"],
                ),
            )
            usina_id = cur.lastrowid

            for etapa, inicio, fim in segmentos:
                conn.execute(
                    """INSERT INTO historico_etapas (usina_id, etapa, data_entrada, data_saida)
                       VALUES (?,?,?,?)""",
                    (usina_id, etapa, inicio.isoformat(), fim.isoformat() if fim else None),
                )

    print(f"Usinas importadas: {len(timeline)}")
    print(f"  Ativas:      {n_ativas}")
    print(f"  Operação:    {n_operacao}")
    print(f"  Rescindidas: {n_rescindidas}")
    if nao_mapeadas:
        print("\nSeções não reconhecidas (usadas com fallback 'TT Usina'):")
        for s in sorted(nao_mapeadas):
            print(f"  - {s}")


if __name__ == "__main__":
    principal()
