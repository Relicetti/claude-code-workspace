import io
import os
from datetime import date, datetime, time as dtime, timedelta

from flask import (Flask, flash, redirect, render_template, request, send_file,
                   session, url_for)
from werkzeug.security import check_password_hash, generate_password_hash

import db
import email_utils

# cabeçalhos do modelo de planilha de importação em massa (ordem = colunas)
COLUNAS_PLANILHA = [
    "UG", "Nome UFV", "Concessionária", "Carteira", "Executivo",
    "Geração média mensal (kWh)", "Data de assinatura (dd/mm/aaaa)", "Etapa inicial",
]

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "chave-de-desenvolvimento-troque-em-producao")
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024  # upload do banco até 25 MB
db.iniciar_banco()
db.bootstrap_admin(generate_password_hash)

ROTAS_PUBLICAS = {"login", "static"}

HORA_SNAPSHOT = dtime(8, 0)  # toda segunda a partir desse horário


def _segunda_da_semana(d):
    return d - timedelta(days=d.weekday())


def _garantir_snapshot_semanal(conn):
    """Roda a cada request: se já passou das 8h da segunda-feira desta semana
    e ainda não existe um retrato dela, tira um agora. Como o app roda local
    (não fica ligado 24h), essa checagem por demanda garante que a foto seja
    tirada assim que alguém abrir o painel depois desse horário, em vez de
    depender de um agendador rodando no instante exato."""
    agora = datetime.now()
    semana = _segunda_da_semana(agora.date())
    if agora < datetime.combine(semana, HORA_SNAPSHOT):
        return
    db.tirar_snapshot(conn, semana.isoformat(), agora.isoformat(timespec="seconds"))


@app.before_request
def exigir_login():
    if request.endpoint in ROTAS_PUBLICAS or request.endpoint is None:
        return None
    if "usuario_id" not in session:
        return redirect(url_for("login", proximo=request.path))
    with db.conectar() as conn:
        _garantir_snapshot_semanal(conn)
    return None


@app.context_processor
def injetar_usuario():
    return {
        "usuario_logado": session.get("usuario_nome"),
        "usuario_admin": session.get("usuario_admin", False),
    }


def _dias_desde(data_iso, referencia=None):
    if not data_iso:
        return None
    referencia = referencia or date.today()
    d = datetime.strptime(data_iso, "%Y-%m-%d").date()
    return (referencia - d).days


def _parse_numero(valor):
    """Converte texto de número em float, ou None. Aceita formato BR
    (1.234,5) e US (1234.5)."""
    valor = (valor or "").strip()
    if not valor:
        return None
    if "," in valor:  # formato BR: ponto é milhar, vírgula é decimal
        valor = valor.replace(".", "").replace(",", ".")
    try:
        return float(valor)
    except ValueError:
        return None


def _eta_dias(etapa_atual, dias_na_etapa, medias):
    if etapa_atual not in db.ETAPAS:
        return None
    idx = db.ETAPAS.index(etapa_atual)
    restante_etapa_atual = medias.get(etapa_atual)
    if restante_etapa_atual is None:
        return None
    total = max(restante_etapa_atual - dias_na_etapa, 0)
    etapas_opcionais = {"Refazer Rateio", "Consumo de Saldo Acumulado", "Rescisão"}
    for etapa in db.ETAPAS[idx + 1:]:
        if etapa in etapas_opcionais:
            continue  # etapas de exceção, não fazem parte do caminho padrão da maioria
        media = medias.get(etapa)
        if media is None:
            return None  # sem dado histórico suficiente pra projetar
        total += media
    return round(total)


def _usinas_ativas_com_metricas(conn):
    usinas = [dict(u) for u in db.listar_ativas(conn)]
    medias = db.medias_por_etapa(conn)
    ultimas_pend = db.ultimas_pendencias_por_usina(conn)
    qtd_pend = db.contar_pendencias_por_usina(conn)
    hoje = date.today()
    for u in usinas:
        u["dias_na_etapa"] = _dias_desde(u["data_entrada_etapa_atual"], hoje)
        u["dias_desde_assinatura"] = _dias_desde(u["data_assinatura_contrato"], hoje)
        media_etapa = medias.get(u["etapa_atual"])
        u["media_etapa"] = media_etapa
        u["atrasada"] = media_etapa is not None and u["dias_na_etapa"] > media_etapa
        u["eta_dias"] = _eta_dias(u["etapa_atual"], u["dias_na_etapa"], medias)
        u["ultima_pendencia"] = ultimas_pend.get(u["id"])
        u["qtd_pendencias"] = qtd_pend.get(u["id"], 0)
    return usinas, medias


# --- login / conta -----------------------------------------------------

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("login.html", erro=None, proximo=request.args.get("proximo", ""))

    username = request.form.get("username", "").strip()
    senha = request.form.get("senha", "")
    with db.conectar() as conn:
        usuario = db.buscar_usuario_por_username(conn, username)

    if usuario is None or not check_password_hash(usuario["senha_hash"], senha):
        return render_template("login.html", erro="Usuário ou senha inválidos.", proximo=request.form.get("proximo", ""))

    session["usuario_id"] = usuario["id"]
    session["usuario_nome"] = usuario["nome"]
    session["usuario_admin"] = bool(usuario["is_admin"])
    proximo = request.form.get("proximo") or url_for("dashboard")
    return redirect(proximo)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/usuarios/novo", methods=["GET", "POST"])
def novo_usuario():
    if not session.get("usuario_admin"):
        return "Só administradores podem cadastrar usuários.", 403

    if request.method == "GET":
        return render_template("novo_usuario.html", erro=None)

    nome = request.form.get("nome", "").strip()
    email = request.form.get("email", "").strip().lower()
    senha = request.form.get("senha", "")
    tipo = request.form.get("tipo", "operador")

    if not nome or "@" not in email or len(senha) < 4:
        return render_template("novo_usuario.html", erro="Preencha nome, e-mail válido e uma senha com pelo menos 4 caracteres.")

    username = email.split("@")[0]

    with db.conectar() as conn:
        if db.buscar_usuario_por_username(conn, username) is not None:
            return render_template("novo_usuario.html", erro=f"Já existe um usuário para o e-mail {email} (usuário '{username}').")
        db.criar_usuario(
            conn, nome, username, generate_password_hash(senha),
            datetime.now().isoformat(timespec="seconds"),
            is_admin=(tipo == "admin"), email=email,
        )
        db.registrar_atividade(
            conn, session["usuario_nome"],
            f"Cadastrou o usuário {nome} ({'admin' if tipo == 'admin' else 'operador'})",
            None, datetime.now().isoformat(timespec="seconds"),
        )

    enviado, motivo = email_utils.enviar_boas_vindas(nome, email, username, senha, request.url_root.rstrip("/"))
    if enviado:
        flash(f"Usuário {nome} criado e e-mail de boas-vindas enviado para {email}.", "success")
    else:
        flash(
            f"Usuário {nome} criado, mas o e-mail não foi enviado ({motivo}). "
            f"Repasse manualmente: usuário '{username}', senha '{senha}'.",
            "warning",
        )
    return redirect(url_for("novo_usuario"))


@app.route("/admin/importar-banco", methods=["GET", "POST"])
def importar_banco():
    if not session.get("usuario_admin"):
        return "Só administradores podem importar o banco.", 403

    if request.method == "GET":
        with db.conectar() as conn:
            qtd_usinas = conn.execute("SELECT COUNT(*) FROM usinas").fetchone()[0]
        return render_template("importar_banco.html", qtd_usinas=qtd_usinas)

    arquivo = request.files.get("banco")
    if not arquivo or not arquivo.filename:
        flash("Selecione um arquivo .db.", "warning")
        return redirect(url_for("importar_banco"))

    # grava o upload na MESMA pasta do banco, pra o os.replace ser atômico e
    # não cruzar drives (no Windows, mover entre C: e D: dá erro).
    destino_tmp = db.DB_PATH + ".importando"
    arquivo.save(destino_tmp)

    ok, msg = db.validar_banco_importado(destino_tmp)
    if not ok:
        os.remove(destino_tmp)
        flash(f"Importação recusada: {msg}", "danger")
        return redirect(url_for("importar_banco"))

    # substitui o banco em uso e reaplica migrações de schema
    os.replace(destino_tmp, db.DB_PATH)
    db.iniciar_banco()
    with db.conectar() as conn:
        db.registrar_atividade(
            conn, session.get("usuario_nome", "?"), "Importou o banco de dados",
            None, datetime.now().isoformat(timespec="seconds"),
        )
    flash(f"Banco importado com sucesso. {msg} Faça login novamente se necessário.", "success")
    return redirect(url_for("dashboard"))


# --- painel --------------------------------------------------------------

@app.route("/")
def dashboard():
    pessoa = request.args.get("pessoa", "").strip()

    with db.conectar() as conn:
        usinas, medias = _usinas_ativas_com_metricas(conn)
        qtd_operacao = db.contar_por_status(conn, "operacao")
        qtd_rescindidas = db.contar_por_status(conn, "rescindida")

        usinas_pessoa = pendencias_pessoa = None
        if pessoa:
            usinas_pessoa = [dict(u) for u in db.buscar_usinas_ativas_por_pessoa(conn, pessoa)]
            hoje = date.today()
            for u in usinas_pessoa:
                u["dias_na_etapa"] = _dias_desde(u["data_entrada_etapa_atual"], hoje)
            pendencias_pessoa = [dict(p) for p in db.buscar_pendencias_por_pessoa(conn, pessoa)]

    def card_etapa(etapa, titulo=None, cor=None):
        idx = db.ETAPAS.index(etapa)
        usinas_etapa = [u for u in usinas if u["etapa_atual"] == etapa]
        return {
            "titulo": titulo or etapa,
            "qtd": len(usinas_etapa),
            "atrasadas": sum(1 for u in usinas_etapa if u["atrasada"]),
            "media": medias.get(etapa),
            "url": url_for("ver_etapa", idx=idx),
            "cor": cor,
            "sub": "usina(s) nesta etapa",
        }

    card_operando = {
        "titulo": "Operando", "qtd": qtd_operacao, "atrasadas": 0, "media": None,
        "url": url_for("ver_operacao"), "cor": "success", "sub": "usina(s) já operando",
    }
    card_rescindidas = {
        "titulo": "Rescindidas", "qtd": qtd_rescindidas, "atrasadas": 0, "media": None,
        "url": url_for("ver_rescindidas"), "cor": "danger", "sub": "usina(s) rescindidas",
    }

    grupos = [
        {"titulo": "Troca de Titularidade", "cards": [
            card_etapa("Assinado c/ Pendência"),
            card_etapa("TT Usina"),
        ]},
        {"titulo": "Rateio", "cards": [
            card_etapa("Sem Clientes"),
            card_etapa("Separando Clientes"),
            card_etapa("TT Cliente"),
            card_etapa("Aguardando Rateio"),
            card_etapa("Aguardando Aprovação Rateio"),
        ]},
        {"titulo": "Operação", "cards": [
            card_operando,
            card_etapa("Refazer Rateio"),
            card_etapa("Consumo de Saldo Acumulado"),
        ]},
        {"titulo": "Rescisão", "cards": [
            card_etapa("Rescisão", titulo="Em Rescisão", cor="warning"),
            card_rescindidas,
        ]},
    ]
    for grupo in grupos:
        grupo["total"] = sum(c["qtd"] for c in grupo["cards"])

    return render_template(
        "dashboard.html",
        grupos=grupos,
        total_ativas=len(usinas),
        qtd_operacao=qtd_operacao,
        qtd_rescindidas=qtd_rescindidas,
        total_geral=len(usinas) + qtd_operacao,
        pessoa=pessoa,
        usinas_pessoa=usinas_pessoa,
        pendencias_pessoa=pendencias_pessoa,
    )


@app.route("/operacao")
def ver_operacao():
    with db.conectar() as conn:
        usinas = [dict(u) for u in db.listar_por_status(conn, "operacao")]
        datas_operacao = db.data_entrada_etapa_final_por_usina(conn, "Operação")

    hoje = date.today()
    for u in usinas:
        u["data_operacao"] = datas_operacao.get(u["id"])
        u["dias_em_operacao"] = _dias_desde(u["data_operacao"], hoje)

    return render_template("operacao.html", usinas=usinas, fluxo_desde_operacao=db.FLUXO_DESDE_OPERACAO)


@app.route("/rescindidas")
def ver_rescindidas():
    with db.conectar() as conn:
        usinas = [dict(u) for u in db.listar_por_status(conn, "rescindida")]
        datas_rescisao = db.data_entrada_etapa_final_por_usina(conn, "Rescindida")

    hoje = date.today()
    for u in usinas:
        u["data_rescisao"] = datas_rescisao.get(u["id"])
        u["dias_desde_rescisao"] = _dias_desde(u["data_rescisao"], hoje)

    return render_template("rescindidas.html", usinas=usinas)


@app.route("/etapa/<int:idx>")
def ver_etapa(idx):
    if idx < 0 or idx >= len(db.ETAPAS):
        return "Etapa inválida", 404
    etapa = db.ETAPAS[idx]

    with db.conectar() as conn:
        usinas, medias = _usinas_ativas_com_metricas(conn)

    usinas_etapa = [u for u in usinas if u["etapa_atual"] == etapa]

    return render_template(
        "etapa.html",
        etapa=etapa,
        idx=idx,
        usinas=usinas_etapa,
        media=medias.get(etapa),
        etapas=db.ETAPAS,
        etapas_finais=db.ETAPAS_FINAIS,
        situacoes=db.SITUACOES,
        total_etapas=len(db.ETAPAS),
        fluxo=db.FLUXO,
    )


@app.route("/comparativo")
def comparativo():
    with db.conectar() as conn:
        semanas = db.semanas_com_snapshot(conn)
        semana_escolhida = request.args.get("semana") or (semanas[0] if semanas else None)
        snapshot = db.snapshot_da_semana(conn, semana_escolhida) if semana_escolhida else {}
        usinas_atuais, _ = _usinas_ativas_com_metricas(conn)

    atuais_por_id = {u["id"]: u for u in usinas_atuais}

    contagem_antes, contagem_agora = {}, {}
    for etapa in db.ETAPAS:
        contagem_antes[etapa] = sum(1 for s in snapshot.values() if s["etapa"] == etapa)
        contagem_agora[etapa] = sum(1 for u in usinas_atuais if u["etapa_atual"] == etapa)

    avancaram, novas, foram_operacao, foram_rescisao, saidas_outras = [], [], [], [], []
    with db.conectar() as conn:
        for usina_id, s in snapshot.items():
            atual = atuais_por_id.get(usina_id)
            if atual is not None:
                if atual["etapa_atual"] != s["etapa"]:
                    avancaram.append({
                        "usina": atual, "etapa_antes": s["etapa"], "etapa_agora": atual["etapa_atual"],
                    })
                continue

            usina_agora = db.buscar_usina(conn, usina_id)
            if usina_agora is None:
                saidas_outras.append({"nome_ufv": s["nome_ufv"], "etapa_antes": s["etapa"], "motivo": "excluída"})
            elif usina_agora["status"] == "operacao":
                foram_operacao.append({"nome_ufv": s["nome_ufv"], "etapa_antes": s["etapa"], "usina_id": usina_id})
            elif usina_agora["status"] == "rescindida":
                foram_rescisao.append({"nome_ufv": s["nome_ufv"], "etapa_antes": s["etapa"], "usina_id": usina_id})
            else:
                saidas_outras.append({"nome_ufv": s["nome_ufv"], "etapa_antes": s["etapa"], "motivo": None})
    for u in usinas_atuais:
        if u["id"] not in snapshot:
            novas.append(u)

    return render_template(
        "comparativo.html",
        semanas=semanas,
        semana_escolhida=semana_escolhida,
        etapas=db.ETAPAS,
        contagem_antes=contagem_antes,
        contagem_agora=contagem_agora,
        avancaram=avancaram,
        novas=novas,
        foram_operacao=foram_operacao,
        foram_rescisao=foram_rescisao,
        saidas_outras=saidas_outras,
        total_antes=len(snapshot),
        total_agora=len(usinas_atuais),
    )


@app.route("/comparativo/tirar-agora", methods=["POST"])
def tirar_snapshot_agora():
    if not session.get("usuario_admin"):
        return "Só administradores podem forçar um novo retrato.", 403
    agora = datetime.now()
    semana = _segunda_da_semana(agora.date())
    with db.conectar() as conn:
        db.tirar_snapshot(conn, semana.isoformat(), agora.isoformat(timespec="seconds"))
    return redirect(url_for("comparativo"))


@app.route("/usina/<int:usina_id>")
def ver_usina(usina_id):
    with db.conectar() as conn:
        usina = db.buscar_usina(conn, usina_id)
        historico = [dict(h) for h in db.historico_usina(conn, usina_id)]
        pendencias = [dict(p) for p in db.pendencias_usina(conn, usina_id)]
        usuarios = db.listar_usuarios(conn)

    if usina is None:
        return "Usina não encontrada", 404

    usina = dict(usina)
    hoje = date.today()
    for h in historico:
        fim = h["data_saida"] or hoje.isoformat()
        h["dias"] = _dias_desde(h["data_entrada"], datetime.strptime(fim, "%Y-%m-%d").date())

    usina["dias_na_etapa"] = _dias_desde(usina["data_entrada_etapa_atual"], hoje)
    usina["dias_desde_assinatura"] = _dias_desde(usina["data_assinatura_contrato"], hoje)

    return render_template(
        "usina.html",
        usina=usina,
        historico=historico,
        pendencias=pendencias,
        usuarios=usuarios,
        etapas=db.ETAPAS,
        etapas_finais=db.ETAPAS_FINAIS,
        situacoes=db.SITUACOES,
        fluxo=db.FLUXO,
        fluxo_desde_operacao=db.FLUXO_DESDE_OPERACAO,
    )


@app.route("/usina/<int:usina_id>/pendencia", methods=["POST"])
def adicionar_pendencia(usina_id):
    texto = request.form.get("texto", "").strip()
    responsavel = request.form.get("responsavel", "").strip()
    if texto:
        with db.conectar() as conn:
            db.adicionar_pendencia(
                conn, usina_id, texto, responsavel, session["usuario_nome"],
                datetime.now().isoformat(timespec="seconds"),
            )
    return redirect(url_for("ver_usina", usina_id=usina_id))


@app.route("/pendencia/<int:pendencia_id>/concluir", methods=["POST"])
def concluir_pendencia(pendencia_id):
    concluir = request.form.get("concluir") == "1"
    with db.conectar() as conn:
        # reabrir (tirar o check) é só pra admin; concluir qualquer um pode
        if not concluir and not session.get("usuario_admin"):
            row = conn.execute("SELECT usina_id FROM pendencias WHERE id = ?", (pendencia_id,)).fetchone()
            flash("Só administradores podem reabrir uma pendência concluída.", "warning")
            return redirect(url_for("ver_usina", usina_id=row["usina_id"]) if row else url_for("dashboard"))
        usina_id = db.marcar_pendencia(
            conn, pendencia_id, concluir, session["usuario_nome"],
            datetime.now().isoformat(timespec="seconds"),
        )
    if usina_id:
        return redirect(url_for("ver_usina", usina_id=usina_id))
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/usina/<int:usina_id>/mudar-etapa", methods=["POST"])
def mudar_etapa(usina_id):
    nova_etapa = request.form.get("nova_etapa", "").strip()
    hoje_iso = date.today().isoformat()
    with db.conectar() as conn:
        usina = db.buscar_usina(conn, usina_id)
        if usina is None:
            return "Usina não encontrada", 404

        destinos_validos = db.FLUXO.get(usina["etapa_atual"], [])
        if nova_etapa not in destinos_validos:
            flash(
                f"Não dá pra mudar {usina['nome_ufv']} de \"{usina['etapa_atual']}\" direto pra "
                f"\"{nova_etapa}\". O fluxo permite ir só pra: {', '.join(destinos_validos) or 'nenhuma (etapa final)'}.",
                "danger",
            )
            return redirect(url_for("ver_usina", usina_id=usina_id))

        if db.tem_pendencia_aberta(conn, usina_id):
            flash(
                f"Não dá pra mudar a etapa de {usina['nome_ufv']}: há pendência(s) em aberto. "
                f"Marque todas como concluídas primeiro.", "warning",
            )
            return redirect(url_for("ver_usina", usina_id=usina_id))
        db.mudar_etapa(conn, usina_id, nova_etapa, hoje_iso, session["usuario_nome"])
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/usina/<int:usina_id>/reabrir-pipeline", methods=["POST"])
def reabrir_pipeline(usina_id):
    nova_etapa = request.form.get("nova_etapa", "").strip()
    if nova_etapa not in db.FLUXO_DESDE_OPERACAO:
        return "Etapa inválida", 400
    hoje_iso = date.today().isoformat()
    with db.conectar() as conn:
        usina = db.buscar_usina(conn, usina_id)
        if usina is None:
            return "Usina não encontrada", 404
        if usina["status"] != "operacao":
            return "Essa usina não está em operação.", 400
        db.reabrir_para_pipeline(conn, usina_id, nova_etapa, hoje_iso, session["usuario_nome"])
        db.registrar_atividade(
            conn, session["usuario_nome"], f"Reabriu do Operando para \"{nova_etapa}\"",
            usina["nome_ufv"], datetime.now().isoformat(timespec="seconds"),
        )
    return redirect(url_for("ver_usina", usina_id=usina_id))


@app.route("/usina/<int:usina_id>/situacao", methods=["POST"])
def mudar_situacao(usina_id):
    situacao = request.form.get("situacao", "").strip()
    if situacao not in db.SITUACOES:
        return "Situação inválida", 400
    agora = datetime.now().isoformat(timespec="seconds")
    with db.conectar() as conn:
        db.mudar_situacao(conn, usina_id, situacao, session["usuario_nome"], agora)
        usina = db.buscar_usina(conn, usina_id)
        db.registrar_atividade(
            conn, session["usuario_nome"], f"Situação da etapa → \"{situacao}\"",
            usina["nome_ufv"] if usina else None, agora,
        )
    return redirect(request.referrer or url_for("ver_usina", usina_id=usina_id))


@app.route("/usinas/nova", methods=["GET", "POST"])
def nova_usina():
    if request.method == "GET":
        with db.conectar() as conn:
            usuarios = db.listar_usuarios(conn)
        return render_template("nova_usina.html", etapas=db.ETAPAS, usuarios=usuarios)

    hoje_iso = date.today().isoformat()
    data_assinatura = request.form.get("data_assinatura", "").strip() or None
    pendencia = request.form.get("pendencia", "").strip()
    etapa_inicial = "Assinado c/ Pendência" if pendencia else "TT Usina"
    with db.conectar() as conn:
        usina_id = db.criar_usina(
            conn,
            ug_raw=request.form.get("ug", "").strip(),
            nome_ufv=request.form.get("nome_ufv", "").strip(),
            concessionaria=request.form.get("concessionaria", "").strip(),
            dono_carteira=request.form.get("dono_carteira", "").strip(),
            executivo=request.form.get("executivo", "").strip(),
            geracao_media_mensal=_parse_numero(request.form.get("geracao_media_mensal")),
            data_assinatura=data_assinatura,
            etapa_inicial=etapa_inicial,
            hoje_iso=hoje_iso,
            pendencia=pendencia,
            responsavel=request.form.get("responsavel", "").strip(),
            autor=session["usuario_nome"],
            criado_em=datetime.now().isoformat(timespec="seconds"),
        )
        db.registrar_atividade(
            conn, session["usuario_nome"], "Cadastrou a usina",
            request.form.get("nome_ufv", "").strip(),
            datetime.now().isoformat(timespec="seconds"),
        )
    return redirect(url_for("ver_usina", usina_id=usina_id))


@app.route("/usinas/modelo")
def modelo_planilha():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Usinas"
    ws.append(COLUNAS_PLANILHA)
    for cel in ws[1]:
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor="1A3A5C")
    # linha de exemplo pra guiar o preenchimento
    ws.append(["12345678", "UFV EXEMPLO", "ENEL RJ", "", "", "12500",
               "15/01/2026", "TT Usina"])
    larguras = [16, 28, 20, 16, 16, 24, 26, 24]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # aba com as etapas válidas, pra consulta
    ws2 = wb.create_sheet("Etapas válidas")
    ws2.append(["Use um destes valores na coluna 'Etapa inicial':"])
    for e in db.ETAPAS:
        ws2.append([e])
    ws2.column_dimensions["A"].width = 40

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer, as_attachment=True, download_name="modelo_usinas.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _valor_celula(v):
    return "" if v is None else str(v).strip()


def _data_celula(v):
    """Converte a célula de data (date do Excel ou texto dd/mm/aaaa) em ISO."""
    if v is None or str(v).strip() == "":
        return None
    if hasattr(v, "date"):  # datetime do openpyxl
        return v.date().isoformat()
    txt = str(v).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(txt, fmt).date().isoformat()
        except ValueError:
            continue
    return None


@app.route("/usinas/importar", methods=["GET", "POST"])
def importar_usinas():
    if request.method == "GET":
        return render_template("importar_usinas.html", colunas=COLUNAS_PLANILHA)

    arquivo = request.files.get("planilha")
    if not arquivo or not arquivo.filename:
        flash("Selecione uma planilha .xlsx.", "warning")
        return redirect(url_for("importar_usinas"))

    from openpyxl import load_workbook
    try:
        wb = load_workbook(arquivo, data_only=True)
    except Exception as e:
        flash(f"Não consegui ler a planilha: {e}", "danger")
        return redirect(url_for("importar_usinas"))

    ws = wb["Usinas"] if "Usinas" in wb.sheetnames else wb.active

    # mapa de coluna por cabeçalho (aceita a planilha na ordem do modelo)
    header = [(_valor_celula(c.value)) for c in ws[1]]
    idx = {nome: header.index(nome) for nome in COLUNAS_PLANILHA if nome in header}
    if "UG" not in idx or "Nome UFV" not in idx:
        flash("A planilha precisa ter pelo menos as colunas 'UG' e 'Nome UFV' (use o modelo).", "danger")
        return redirect(url_for("importar_usinas"))

    def cel(row, nome):
        return row[idx[nome]] if nome in idx and idx[nome] < len(row) else None

    hoje_iso = date.today().isoformat()
    criadas = ignoradas = erros = 0
    with db.conectar() as conn:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            ug = _valor_celula(cel(row, "UG"))
            nome = _valor_celula(cel(row, "Nome UFV"))
            if not ug and not nome:
                continue  # linha vazia
            if not ug or not nome:
                erros += 1
                continue
            if db.ug_existe(conn, ug):
                ignoradas += 1
                continue
            etapa = _valor_celula(cel(row, "Etapa inicial"))
            if etapa not in db.ETAPAS:
                etapa = "TT Usina"
            db.criar_usina(
                conn,
                ug_raw=ug,
                nome_ufv=nome,
                concessionaria=_valor_celula(cel(row, "Concessionária")),
                dono_carteira=_valor_celula(cel(row, "Carteira")),
                executivo=_valor_celula(cel(row, "Executivo")),
                geracao_media_mensal=_parse_numero(_valor_celula(cel(row, "Geração média mensal (kWh)"))),
                data_assinatura=_data_celula(cel(row, "Data de assinatura (dd/mm/aaaa)")),
                etapa_inicial=etapa,
                hoje_iso=hoje_iso,
                pendencia="",
                responsavel="",
                autor=session["usuario_nome"],
                criado_em=datetime.now().isoformat(timespec="seconds"),
            )
            criadas += 1
        if criadas:
            db.registrar_atividade(
                conn, session["usuario_nome"],
                f"Importou {criadas} usina(s) via planilha", None,
                datetime.now().isoformat(timespec="seconds"),
            )

    partes = [f"{criadas} usina(s) criada(s)"]
    if ignoradas:
        partes.append(f"{ignoradas} ignorada(s) (UG já existia)")
    if erros:
        partes.append(f"{erros} linha(s) com erro (faltando UG ou nome)")
    flash("Importação concluída: " + " · ".join(partes) + ".", "success" if criadas else "warning")
    return redirect(url_for("dashboard"))


@app.route("/usina/<int:usina_id>/editar", methods=["GET", "POST"])
def editar_usina(usina_id):
    if not session.get("usuario_admin"):
        return "Só administradores podem editar os dados da usina.", 403

    with db.conectar() as conn:
        usina = db.buscar_usina(conn, usina_id)
        if usina is None:
            return "Usina não encontrada", 404

        if request.method == "GET":
            usuarios = db.listar_usuarios(conn)
            return render_template("editar_usina.html", usina=dict(usina), usuarios=usuarios)

        db.atualizar_usina(conn, usina_id, {
            "ug_raw": request.form.get("ug", "").strip(),
            "nome_ufv": request.form.get("nome_ufv", "").strip(),
            "concessionaria": request.form.get("concessionaria", "").strip(),
            "dono_carteira": request.form.get("dono_carteira", "").strip(),
            "executivo": request.form.get("executivo", "").strip(),
            "geracao_media_mensal": _parse_numero(request.form.get("geracao_media_mensal")),
            "data_assinatura_contrato": request.form.get("data_assinatura", "").strip() or None,
        })
        db.registrar_atividade(
            conn, session["usuario_nome"], "Editou os dados da usina",
            request.form.get("nome_ufv", "").strip(),
            datetime.now().isoformat(timespec="seconds"),
        )
    return redirect(url_for("ver_usina", usina_id=usina_id))


@app.route("/usina/<int:usina_id>/excluir", methods=["POST"])
def excluir_usina(usina_id):
    if not session.get("usuario_admin"):
        return "Só administradores podem excluir usinas.", 403
    with db.conectar() as conn:
        usina = db.buscar_usina(conn, usina_id)
        nome = usina["nome_ufv"] if usina else f"id {usina_id}"
        db.excluir_usina(conn, usina_id)
        db.registrar_atividade(
            conn, session["usuario_nome"], "Excluiu a usina", nome,
            datetime.now().isoformat(timespec="seconds"),
        )
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/admin/log")
def ver_log():
    if not session.get("usuario_admin"):
        return "Só administradores podem ver o log.", 403
    with db.conectar() as conn:
        atividades = db.listar_atividades(conn)
    for a in atividades:
        a["quando_fmt"] = (a["quando"] or "").replace("T", " ")
    return render_template("log.html", atividades=atividades)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5010))
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)
