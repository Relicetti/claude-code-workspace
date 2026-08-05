import io
import os

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font

import geracao

app = Flask(__name__)


@app.route("/")
def index():
    return render_template("index.html", estados=geracao.listar_estados())


@app.route("/cidades")
def cidades():
    uf = request.args.get("uf", "")
    return jsonify([c["nome"] for c in geracao.listar_cidades(uf)])


def _ler_parametros(dados):
    uf = dados.get("uf", "")
    cidade = dados.get("cidade", "")
    potencia_cc = float(str(dados["potencia_cc"]).replace(",", "."))
    potencia_ca = float(str(dados["potencia_ca"]).replace(",", "."))
    return uf, cidade, potencia_cc, potencia_ca


@app.route("/calcular", methods=["POST"])
def calcular():
    try:
        uf, cidade, potencia_cc, potencia_ca = _ler_parametros(request.form)
    except (KeyError, ValueError):
        return render_template("index.html", estados=geracao.listar_estados(),
                                erro="Informe Potência CC e Potência CA válidas.")

    coords = geracao.buscar_coordenadas(uf, cidade)
    if not coords:
        return render_template("index.html", estados=geracao.listar_estados(),
                                erro=f"Cidade '{cidade}' não encontrada em {uf}.")
    lat, lon = coords

    try:
        resultado = geracao.calcular_geracao(lat, lon, potencia_cc, potencia_ca)
    except Exception as e:
        return render_template("index.html", estados=geracao.listar_estados(),
                                erro=f"Erro ao consultar o PVGIS: {e}")

    return render_template(
        "resultado.html",
        cidade=cidade, uf=uf, lat=lat, lon=lon,
        potencia_cc=potencia_cc, potencia_ca=potencia_ca,
        r=resultado,
    )


@app.route("/exportar-excel")
def exportar_excel():
    uf, cidade, potencia_cc, potencia_ca = _ler_parametros(request.args)

    coords = geracao.buscar_coordenadas(uf, cidade)
    if not coords:
        return f"Cidade '{cidade}' não encontrada em {uf}.", 404
    lat, lon = coords
    resultado = geracao.calcular_geracao(lat, lon, potencia_cc, potencia_ca)

    wb = Workbook()
    ws = wb.active
    ws.title = "Geração CA"

    ws.append([f"Estimativa de Geração — {cidade}/{uf}"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"Potência CC: {potencia_cc:g} kWp   |   Potência CA: {potencia_ca:g} kW"])
    ws.append([])

    cabecalho = ["Mês", "Geração CA (kWh)", "FC do mês (%)"]
    ws.append(cabecalho)
    for cel in ws[4]:
        cel.font = Font(bold=True)

    for m in resultado["meses"]:
        ws.append([m["mes"], round(m["energia_ca_kwh"], 1), round(m["fc"], 1)])

    ws.append(["Total anual", round(resultado["energia_ca_total_kwh"], 1), None])
    for cel in ws[ws.max_row]:
        cel.font = Font(bold=True)

    ws.append(["Média mensal", round(resultado["energia_ca_media_mensal_kwh"], 1), round(resultado["fc_medio_anual"], 1)])
    for cel in ws[ws.max_row]:
        cel.font = Font(bold=True)

    for col, largura in zip("ABC", (18, 20, 16)):
        ws.column_dimensions[col].width = largura

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nome_arquivo = f"geracao_{cidade}_{uf}.xlsx".replace(" ", "_")
    return send_file(
        buffer,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5003))
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)
